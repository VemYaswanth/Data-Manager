from openpyxl import load_workbook
import io

def extract_text_from_xlsx(raw_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
    text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            text.append(" ".join([str(v) for v in row if v is not None]))
    return "\n".join(text)
